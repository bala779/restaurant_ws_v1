from ament_index_python.packages import get_package_share_directory
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pathlib import Path
import json
import time
import os

package_share = get_package_share_directory('butler_order_manager')
ORDERS_JSON = Path(package_share)/'reference'/'orders.json'
EXCEL_FILE = "orders_log.xlsx"


print(ORDERS_JSON)

DEFAULT_DATA = {
    "current_order_table": '',
    "order_list": []
}

def log_to_excel(data: dict):
    try:
        # If file exists, load it, else create new
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Timestamp", "    Table", "    Items", "    Goal Status", "    Result", "    Message"]) 
            for cell in ws[1]:
                cell.font = Font(bold=True)

        ws.append([
            time.strftime("%Y-%m-%d %H:%M:%S"),
            data.get("table", ""),
            data.get("items", ""),
            data.get("goal_status", ""),
            data.get("result", ""),
            data.get("message", "")
        ])
        wb.save(EXCEL_FILE)
    except Exception as e:
        log_print(f"Excel Logging Error: {e}")


def _ensure_file():
    """Ensure JSON file exists with default structure"""
    try:
        os.makedirs(os.path.dirname(ORDERS_JSON), exist_ok=True)
        if not os.path.exists(ORDERS_JSON):
            write_json(DEFAULT_DATA)
    except Exception as e:
        log_print(f"Utils: Error ensuring file: {e}")


def read_json():
    """Read orders.json and return data"""
    try:
        _ensure_file()
        with open(ORDERS_JSON, "r") as f:
            return json.load(f)
    except Exception as e:
        log_print(f"Utils: Error reading JSON file: {e}")
        return DEFAULT_DATA.copy()


def write_json(data):
    """Write data back to orders.json"""
    try:
        with open(ORDERS_JSON, "w") as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        log_print(f"Utils: Error writing JSON file: {e}")


# --------------------------
# Order Management Functions
# --------------------------

def remove_order(table_name: str, starts=0):
    try:
        data = read_json()
        for i in range(starts, len(data.get("order_list", []))):
            order = data["order_list"][i]
            if order.get("table") == table_name:
                data["order_list"].remove(order)
                break
        write_json(data)
    except Exception as e:
        log_print(f"Utils: Error removing order for {table_name}: {e}")


def get_current_order_table():
    try:
        data = read_json()
        table = data.get("current_order_table")
        log_print(f"Utils: Get Current Order : {table}")
        return table
    except Exception as e:
        log_print(f"Utils: Error getting current order table: {e}")
        return ''


def set_current_order_table(table_name: str):
    try:
        data = read_json()
        data["current_order_table"] = table_name
        log_print(f"Utils: Set Current Order : {table_name}")
        write_json(data)
    except Exception as e:
        log_print(f"Utils: Error setting current order table: {e}")


def clear_current_order_table():
    try:
        data = read_json()
        data["current_order_table"] = ''
        write_json(data)
    except Exception as e:
        log_print(f"Utils: Error clearing current order table: {e}")


def get_order_list():
    try:
        data = read_json()
        return data.get("order_list", [])
    except Exception as e:
        log_print(f"Utils: Error getting order list: {e}")
        return []


def add_order(order: dict):
    try:
        data = read_json()
        data.setdefault("order_list", []).append(order)
        write_json(data)
    except Exception as e:
        log_print(f"Utils: Error adding order: {e}")


def clear_order_list():
    try:
        data = read_json()
        data["order_list"] = []
        write_json(data)
    except Exception as e:
        log_print(f"Utils: Error clearing order list: {e}")


def clear_all_caches():
    try:
        write_json(DEFAULT_DATA.copy())
    except Exception as e:
        log_print(f"Utils: Error clearing all caches: {e}")


def log_print(msg):
    try:
        print(msg, flush=True)
    except Exception:
        pass
